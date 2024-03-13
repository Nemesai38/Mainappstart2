import io
import tkinter
import os
import sys
import sqlite3
import pathlib
import openpyxl
from tkinter import ttk
from tkinter import *
from PIL import ImageTk, Image
from tkinter import messagebox
from tkinter import filedialog
from openpyxl import Workbook


# Main Window
window = tkinter.Tk()
window.title("Personnel Entry Form")
window.geometry("1250x680")
window.resizable(False, False)

# Main Frame
frame = tkinter.Frame(window)
frame.pack()

# define i variable for indexing fetched db results
i = 0

# Excel file creation
excel_file = pathlib.Path('Personnel_Info.xlsx')
if excel_file.exists():
    pass
else:
    excel_file = Workbook()
    sheet = excel_file.active
    sheet['A1'] = "Serial"
    sheet['B1'] = "Army Number"
    sheet['C1'] = "Rank"
    sheet['D1'] = "First Name"
    sheet['E1'] = "Last Name"
    sheet['F1'] = "Gender"
    sheet['G1'] = "Unit"
    sheet['H1'] = "Corps"
    sheet['I1'] = "Appointment"
    sheet['J1'] = "Deployment"
    sheet['K1'] = "State"
    sheet['L1'] = "Trade"
    sheet['M1'] = "Address"
    sheet['N1'] = "Marital Status"

    excel_file.save(r'Personnel_Info.xlsx')


# Function to select image file from os
# and display on form selected image
def upload_img():
    global img_label
    global filename
    filename = filedialog.askopenfilename(initialdir=os.getcwd(), title="Select Photo File", filetypes=(("JPG File", "*.jpg"),
                                                                                                        ("PNG File", "*.png"),
                                                                                                        ("All Files", "*.txt")))
    img_label = Label(per_img_frame, image=photo)
    img_label.place(x=0, y=0)
    try:
        img_name = Image.open(filename)
        size = (300, 310)
        resized_photo = img_name.resize(size)
        resized_photo.save(filename)
        photo2 = ImageTk.PhotoImage(resized_photo)
        img_label.config(image=photo2)
        img_label = photo2
    except AttributeError:
        pass


def clear_image():
    global img_label
    img_label = Label(per_img_frame, image=photo)
    img_label.place(x=0, y=0)


# Function to collect personnel information into database
# and create one if one doesn't exist
def reg_per():
    if army_no_entry.get() == "" or rank_entry.get() == "" or first_name_entry.get() == "" or last_name_entry.get()\
       == "" or unit_entry.get() == "" or corps_entry.get() == "" or appointment_entry.get() == "" or deploy_entry.get()\
       == "" or state_entry.get() == "" or trade_entry.get() == "" or address_entry.get("1.0", 'end') == "":
        tkinter.messagebox.showinfo(title="ERROR", message="Please all fields are required to be filled")
    elif Gender_radio.get() == "":
        tkinter.messagebox.showerror(title="Error", message="Please select Gender")
    elif Marital_Status_radio.get() == "":
        tkinter.messagebox.showerror(title="Error", message="Please select Marital Status")
    else:
        # User Info
        try:
            with open(filename, 'rb') as file:
                photo_image = file.read()
        except NameError:
            pass
        army_number = army_no_entry.get()
        rank = rank_entry.get()
        firstname = first_name_entry.get()
        lastname = last_name_entry.get()
        sex = Gender_radio.get()
        unit = unit_entry.get()
        corps = corps_entry.get()
        appointment = appointment_entry.get()
        deployment = deploy_entry.get()
        state = state_entry.get()
        trade = trade_entry.get()
        address = address_entry.get("1.0", 'end-1c')
        marital_status = Marital_Status_radio.get()
        try:
            image = photo_image
        except NameError:
            image = ""

        # Create Table
        conn = sqlite3.connect('data.db')
        table_create_query = '''CREATE TABLE IF NOT EXISTS Personnel_Info (serial INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL, 
        armynumber TEXT, rank TEXT, firstname TEXT, lastname TEXT, sex TEXT, unit TEXT, corps TEXT, appointment TEXT, 
        deployment TEXT, state TEXT, trade TEXT, address TEXT, maritalstatus TEXT, image BLOB)'''
        
        conn.execute(table_create_query)

        # Insert Data
        data_insert_query = '''INSERT INTO Personnel_Info (armynumber, rank, firstname, lastname, sex, unit, corps, appointment,
        deployment, state, trade, address, maritalstatus, image) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)'''
        data_insert_tuple = (army_number, rank, firstname, lastname, sex, unit, corps, appointment, deployment, state, trade,
                             address, marital_status, image,)
        cursor = conn.cursor()
        cursor.execute(data_insert_query, data_insert_tuple)
        conn.commit()
        cursor.close()
        conn.close()

        # entry into excel file
        file = openpyxl.load_workbook('Personnel_Info.xlsx')
        sheets = file.active
        sheets.cell(column=1, row=sheets.max_row+1, value=sheets.max_row+1)
        sheets.cell(column=2, row=sheets.max_row, value=army_number)
        sheets.cell(column=3, row=sheets.max_row, value=rank)
        sheets.cell(column=4, row=sheets.max_row, value=firstname)
        sheets.cell(column=5, row=sheets.max_row, value=lastname)
        sheets.cell(column=6, row=sheets.max_row, value=sex)
        sheets.cell(column=7, row=sheets.max_row, value=unit)
        sheets.cell(column=8, row=sheets.max_row, value=corps)
        sheets.cell(column=9, row=sheets.max_row, value=appointment)
        sheets.cell(column=10, row=sheets.max_row, value=deployment)
        sheets.cell(column=11, row=sheets.max_row, value=state)
        sheets.cell(column=12, row=sheets.max_row, value=trade)
        sheets.cell(column=13, row=sheets.max_row, value=address)
        sheets.cell(column=14, row=sheets.max_row, value=marital_status)
        file.save(r'Personnel_Info.xlsx')

        # Notify info entry success
        tkinter.messagebox.showinfo(title="DONE", message="Personnel has been registered successfully")

        # Clear fields
        clear_all_fields()
        return


# Function to clear all input fields at will
def clear_all_fields():
    global img_label
    Serial_Number.set(0)
    army_no_entry.delete(0, 'end')
    rank_entry.set('')
    first_name_entry.delete(0, 'end')
    last_name_entry.delete(0, 'end')
    unit_entry.delete(0, 'end')
    corps_entry.delete(0, 'end')
    appointment_entry.delete(0, 'end')
    deploy_entry.delete(0, 'end')
    state_entry.set('')
    trade_entry.delete(0, 'end')
    address_entry.delete('1.0', 'end')
    img_label = Label(per_img_frame, image=photo)
    img_label.place(x=0, y=0)
    return


# Function to return to DB Hub (close dialog box at the moment)
def go_back():
    window.destroy()
    return


# Function to set form to initial state
def restart():
    python = sys.executable
    os.execl(python, python, * sys.argv)


# ################### Command to start to view database on form ##################
def peruse_database():
    global img_label
    global i
    global blob_image
    clear_all_fields()
    go_back_btn.config(state='normal')
    go_forward_btn.config(state='normal')
    upload_image_btn.config(state='normal')
    clear_image_btn.config(state='normal')
    clear_button.config(state='disabled')
    register_personnel_btn.config(state='disabled')
    try:
        search_next_btn.config(state='disabled')
        search_prev_btn.config(state='disabled')
    except NameError:
        pass
    conn = sqlite3.connect('data.db')
    cursor = conn.cursor()
    c = cursor.execute("SELECT * FROM Personnel_Info ORDER BY CASE rank WHEN 'Brig Gen' THEN 0 WHEN 'Col' THEN 1 WHEN 'Lt Col' THEN 2 \
                       WHEN 'Maj' THEN 3 WHEN 'Capt' THEN 4 WHEN 'Lt' THEN 5 WHEN '2Lt' THEN 6 WHEN 'AWO' THEN 7 WHEN 'MWO' THEN 8 \
                       WHEN 'WO' THEN 9 WHEN 'SSgt' THEN 10 WHEN 'Sgt' THEN 11 WHEN 'Cpl' THEN 12 WHEN 'LCpl' THEN 13 \
                       WHEN 'Pte'  THEN 14 END, armynumber ASC")
    r = c.fetchall()

    Serial_Number.set(r[i][0])
    Army_Number.set(r[i][1])
    Rank.set(r[i][2])
    First_Name.set(r[i][3])
    Last_Name.set(r[i][4])
    Gender_radio.set(r[i][5])
    Unit.set(r[i][6])
    Corps.set(r[i][7])
    Appointment.set(r[i][8])
    Deployment.set(r[i][9])
    State.set(r[i][10])
    Trade.set(r[i][11])
    address_entry.insert(INSERT, r[i][12])
    Marital_Status_radio.set(r[i][13])
    try:
        blob_image = r[i][14]
        blob = io.BytesIO(blob_image)
        image = Image.open(blob)
        per_image = ImageTk.PhotoImage(image)
        img_label.config(image=per_image)
        img_label = per_image
    except TypeError:
        pass

    cursor.close()
    conn.close()
    return


# function to get max rows for scrolling through database
def get_max_rows():
    conn = sqlite3.connect('data.db')
    cursor = conn.cursor()
    c = cursor.execute("SELECT COUNT(*) FROM Personnel_Info")
    result = c.fetchone()
    total_row = result[0]
    cursor.close()
    conn.close()
    return total_row


# #################### functions to scroll through whole database results #############################
def next_personnel():
    global i
    if Army_Number.get() == "":
        return
    if i == get_max_rows() - 1:
        tkinter.messagebox.showinfo("End", "Last Personnel in DB")
    else:
        i += 1
        peruse_database()
    return


def previous_personnel():
    global i
    if Army_Number.get() == "":
        return 
    if i == 0:
        tkinter.messagebox.showinfo("Begin", "First Personnel in DB")
    else:
        i -= 1
        peruse_database()
    return


# ##################### Update Record ###################################
def update_record():
    if army_no_entry.get() == "" or rank_entry.get() == "" or first_name_entry.get() == "" or last_name_entry.get()\
       == "" or unit_entry.get() == "" or corps_entry.get() == "" or appointment_entry.get() == "" or deploy_entry.get()\
       == "" or state_entry.get() == "" or trade_entry.get() == "" or address_entry.get("1.0", 'end') == "":
        tkinter.messagebox.showinfo(title="ERROR", message="Please all fields are required to be filled")
    elif Gender_radio.get() == "":
        tkinter.messagebox.showerror(title="Error", message="Please select Gender")
    elif Marital_Status_radio.get() == "":
        tkinter.messagebox.showerror(title="Error", message="Please select Marital Status")
    else:
        try:
            with open(filename, 'rb') as file:
                photo_image = file.read()
        except NameError:
            pass
        serial = Serial_Number.get()
        army_number = Army_Number.get()
        rank = Rank.get()
        firstname = First_Name.get()
        lastname = Last_Name.get()
        sex = Gender_radio.get()
        unit = Unit.get()
        corps = Corps.get()
        appointment = Appointment.get()
        deployment = Deployment.get()
        state = State.get()
        trade = Trade.get()
        address = address_entry.get("1.0", 'end-1c')
        marital_status = Marital_Status_radio.get()
        try:
            image = photo_image
        except NameError:
            image = blob_image
        except Exception as e:
            image = ""

        # Connect to database
        conn = sqlite3.connect('data.db')

        # Edit Data
        data_update_query = ("UPDATE Personnel_Info SET armynumber = ?, rank = ?, firstname = ?, lastname = ?, sex = ?, unit = ?,\
                       corps = ?, appointment = ?, deployment = ?, state = ?, trade = ?, address = ?, maritalstatus = ?,\
                       image = ? WHERE serial = ?")
        data_update_tuple = (army_number, rank, firstname, lastname, sex, unit, corps, appointment, deployment, state,
                             trade, address, marital_status, image, serial)
        cursor = conn.cursor()
        cursor.execute(data_update_query, data_update_tuple)
        conn.commit()
        cursor.close()
        conn.close()

        # Entry into Excel File
        file = openpyxl.load_workbook('Personnel_Info.xlsx')
        sheets = file.active

        column_b = sheets['B']

        for cell in column_b:
            if cell.value == army_number:
                row = cell.row

        sheets.cell(column=1, row=row, value=serial)
        sheets.cell(column=2, row=row, value=army_number)
        sheets.cell(column=3, row=row, value=rank)
        sheets.cell(column=4, row=row, value=firstname)
        sheets.cell(column=5, row=row, value=lastname)
        sheets.cell(column=6, row=row, value=sex)
        sheets.cell(column=7, row=row, value=unit)
        sheets.cell(column=8, row=row, value=corps)
        sheets.cell(column=9, row=row, value=appointment)
        sheets.cell(column=10, row=row, value=deployment)
        sheets.cell(column=11, row=row, value=state)
        sheets.cell(column=12, row=row, value=trade)
        sheets.cell(column=13, row=row, value=address)
        sheets.cell(column=14, row=row, value=marital_status)
        file.save(r'Personnel_Info.xlsx')

        # Notify info entry success
        tkinter.messagebox.showinfo(title="DONE", message="Personnel record updated successfully")

        # Clear fields
        clear_all_fields()
        return


# function to delete record from database
def delete_record():
    if Serial_Number.get() == 0:
        tkinter.messagebox.showerror("Error", "First Select a Personnel")
    else:
        decision = tkinter.messagebox.askyesno("Confirm Removal", "Personnel information will not be recoverable!\nClick "
                                               "Yes to confirm deletion")
        if decision:
            del_serial = Serial_Number.get()
            connection = sqlite3.connect('data.db')
            cursor = connection.cursor()
            cursor.execute("DELETE FROM Personnel_Info WHERE serial = ?", (f'{del_serial}',))
            connection.commit()
            cursor.close()
            connection.close()
            tkinter.messagebox.showinfo("Done", "Personnel deleted Successfully!")
            clear_all_fields()
            reset_i()
        else:
            pass
        return


# function to reset variable i to avoid out of range Error when
# reestablishing connection or change of database query
def reset_i():
    global i
    if i != 0:
        i = 0
    return


# function to input search result in form
def input_function(r):
    i = 0
    global img_label
    global blob_image
    clear_all_fields()
    upload_image_btn.config(state='disabled')
    clear_image_btn.config(state='disabled')
    clear_button.config(state='disabled')
    register_personnel_btn.config(state='disabled')
    Serial_Number.set(r[i][0])
    Army_Number.set(r[i][1])
    Rank.set(r[i][2])
    First_Name.set(r[i][3])
    Last_Name.set(r[i][4])
    Gender_radio.set(r[i][5])
    Unit.set(r[i][6])
    Corps.set(r[i][7])
    Appointment.set(r[i][8])
    Deployment.set(r[i][9])
    State.set(r[i][10])
    Trade.set(r[i][11])
    address_entry.insert(INSERT, r[i][12])
    Marital_Status_radio.set(r[i][13])
    blob_image = r[i][14]
    blob = io.BytesIO(blob_image)
    image = Image.open(blob)
    per_image = ImageTk.PhotoImage(image)
    img_label.config(image=per_image)
    img_label = per_image


# Function to search using entered name string
def search_database_name():
    search_str = search_string_entry.get()
    if search_str == "":
        tkinter.messagebox.showerror("Error", "Please enter name or part of name in FIND field")
    else:
        conn = sqlite3.connect('data.db')
        cursor = conn.cursor()
        c = cursor.execute("SELECT * FROM Personnel_Info WHERE firstname LIKE ? or lastname LIKE ?",
                           (f'%{search_str}%', f'%{search_str}%',))
        r = c.fetchall()
        if r:
            input_function(r)
        else:
            tkinter.messagebox.showinfo("Error", "Name not found in DB")
        conn.close()


# match case function to establish various search queries
def search_master(button):
    global i
    global img_label
    global blob_image
    global search_result
    global search_next_btn
    global search_prev_btn
    search_string = search_string_entry.get()
    if search_string == "":
        tkinter.messagebox.showerror("Error", "Please enter what to search in FIND field")
    else:
        connection = sqlite3.connect("data.db")
        match button:
            case 'search_army_no_btn':
                cursor = connection.cursor()
                army_no_search_result = cursor.execute("SELECT * FROM Personnel_Info WHERE armynumber LIKE ?",
                                                       (f'%{search_string}%',))
                search_result = army_no_search_result.fetchall()
                search_next_btn = tkinter.Button(per_info_frame, text='Next', bg="#93c47d", command=lambda: next_per(
                    'search_army_no_btn'))
                search_next_btn.grid(row=6, column=3, sticky='e')
                search_prev_btn = tkinter.Button(per_info_frame, text='Back ', bg="#93c47d", command=lambda: prev_per(
                    'search_army_no_btn'))
                search_prev_btn.grid(row=6, column=3, sticky='w')
            case 'search_rank_btn':
                cursor = connection.cursor()
                rank_search_result = cursor.execute(f"SELECT * FROM Personnel_Info WHERE rank LIKE ? ORDER BY armynumber ASC",
                                                    (search_string,))
                search_result = rank_search_result.fetchall()
                search_next_btn = tkinter.Button(per_info_frame, text='Next', bg="#93c47d", command=lambda: next_per(
                    'search_rank_btn'))
                search_next_btn.grid(row=6, column=3, sticky='e')
                search_prev_btn = tkinter.Button(per_info_frame, text='Back ', bg="#93c47d", command=lambda: prev_per(
                    'search_rank_btn'))
                search_prev_btn.grid(row=6, column=3, sticky='w')
            case 'search_deployment_btn':
                cursor = connection.cursor()
                deployment_search_result = cursor.execute(f"SELECT * FROM Personnel_Info WHERE deployment LIKE ? ORDER BY CASE \
                                                          rank WHEN 'Brig Gen' THEN 0 WHEN 'Col' THEN 1 WHEN 'Lt Col' THEN 2 \
                       WHEN 'Maj' THEN 3 WHEN 'Capt' THEN 4 WHEN 'Lt' THEN 5 WHEN '2Lt' THEN 6 WHEN 'AWO' THEN 7 WHEN 'MWO' THEN 8 \
                       WHEN 'WO' THEN 9 WHEN 'SSgt' THEN 10 WHEN 'Sgt' THEN 11 WHEN 'Cpl' THEN 12 WHEN 'LCpl' THEN 13 \
                       WHEN 'Pte'  THEN 14 END, armynumber ASC", (search_string,))
                search_result = deployment_search_result.fetchall()
                search_next_btn = tkinter.Button(per_info_frame, text='Next', bg="#93c47d", command=lambda: next_per(
                    'search_deployment_btn'))
                search_next_btn.grid(row=6, column=3, sticky='e')
                search_prev_btn = tkinter.Button(per_info_frame, text='Back ', bg="#93c47d", command=lambda: prev_per(
                    'search_deployment_btn'))
                search_prev_btn.grid(row=6, column=3, sticky='w')
            case 'search_state_btn':
                cursor = connection.cursor()
                state_search_result = cursor.execute("SELECT * FROM Personnel_Info WHERE state LIKE ? ORDER BY CASE \
                                                          rank WHEN 'Brig Gen' THEN 0 WHEN 'Col' THEN 1 WHEN 'Lt Col' THEN 2 \
                       WHEN 'Maj' THEN 3 WHEN 'Capt' THEN 4 WHEN 'Lt' THEN 5 WHEN '2Lt' THEN 6 WHEN 'AWO' THEN 7 WHEN 'MWO' THEN 8 \
                       WHEN 'WO' THEN 9 WHEN 'SSgt' THEN 10 WHEN 'Sgt' THEN 11 WHEN 'Cpl' THEN 12 WHEN 'LCpl' THEN 13 \
                       WHEN 'Pte'  THEN 14 END, armynumber ASC", (search_string,))
                search_result = state_search_result.fetchall()
                search_next_btn = tkinter.Button(per_info_frame, text='Next', bg="#93c47d", command=lambda: next_per(
                    'search_state_btn'))
                search_next_btn.grid(row=6, column=3, sticky='e')
                search_prev_btn = tkinter.Button(per_info_frame, text='Back ', bg="#93c47d", command=lambda: prev_per(
                    'search_state_btn'))
                search_prev_btn.grid(row=6, column=3, sticky='w')
            case 'search_trade_btn':
                cursor = connection.cursor()
                trade_search_result = cursor.execute(f"SELECT * FROM Personnel_Info WHERE trade LIKE ? ORDER BY CASE rank WHEN \
                                                     'Brig Gen' THEN 0 WHEN 'Col' THEN 1 WHEN 'Lt Col' THEN 2 \
                       WHEN 'Maj' THEN 3 WHEN 'Capt' THEN 4 WHEN 'Lt' THEN 5 WHEN '2Lt' THEN 6 WHEN 'AWO' THEN 7 WHEN 'MWO' THEN 8 \
                       WHEN 'WO' THEN 9 WHEN 'SSgt' THEN 10 WHEN 'Sgt' THEN 11 WHEN 'Cpl' THEN 12 WHEN 'LCpl' THEN 13 \
                       WHEN 'Pte'  THEN 14 END, armynumber ASC", (f'%{search_string}%',))
                search_result = trade_search_result.fetchall()
                search_next_btn = tkinter.Button(per_info_frame, text='Next', bg="#93c47d", command=lambda: next_per(
                    'search_trade_btn'))
                search_next_btn.grid(row=6, column=3, sticky='e')
                search_prev_btn = tkinter.Button(per_info_frame, text='Back ', bg="#93c47d", command=lambda: prev_per(
                    'search_trade_btn'))
                search_prev_btn.grid(row=6, column=3, sticky='w')
            case 'search_corps_btn':
                cursor = connection.cursor()
                corps_search_result = cursor.execute(f"SELECT * FROM Personnel_Info WHERE corps LIKE ? ORDER BY CASE rank WHEN \
                                                      'Brig Gen' THEN 0 WHEN 'Col' THEN 1 WHEN 'Lt Col' THEN 2 \
                       WHEN 'Maj' THEN 3 WHEN 'Capt' THEN 4 WHEN 'Lt' THEN 5 WHEN '2Lt' THEN 6 WHEN 'AWO' THEN 7 WHEN 'MWO' THEN 8 \
                       WHEN 'WO' THEN 9 WHEN 'SSgt' THEN 10 WHEN 'Sgt' THEN 11 WHEN 'Cpl' THEN 12 WHEN 'LCpl' THEN 13 \
                       WHEN 'Pte'  THEN 14 END, armynumber ASC", (f'%{search_string}%',))
                search_result = corps_search_result.fetchall()
                search_next_btn = tkinter.Button(per_info_frame, text='Next', bg="#93c47d", command=lambda: next_per(
                    'search_corps_btn'))
                search_next_btn.grid(row=6, column=3, sticky='e')
                search_prev_btn = tkinter.Button(per_info_frame, text='Back ', bg="#93c47d", command=lambda: prev_per(
                    'search_corps_btn'))
                search_prev_btn.grid(row=6, column=3, sticky='w')
            case 'search_appointment_btn':
                cursor = connection.cursor()
                appointment_search_result = cursor.execute("SELECT * FROM Personnel_Info WHERE appointment LIKE ?",
                                                           (f'%{search_string}%',))
                search_result = appointment_search_result.fetchall()
                search_next_btn = tkinter.Button(per_info_frame, text='Next', bg="#93c47d", command=lambda: next_per(
                    'search_appointment_btn'))
                search_next_btn.grid(row=6, column=3, sticky='e')
                search_prev_btn = tkinter.Button(per_info_frame, text='Back ', bg="#93c47d", command=lambda: prev_per(
                    'search_appointment_btn'))
                search_prev_btn.grid(row=6, column=3, sticky='w')
            case 'search_gender_btn':
                cursor = connection.cursor()
                gender_search_result = cursor.execute("SELECT * FROM Personnel_Info WHERE LOWER(sex) = ? ORDER BY CASE rank WHEN \
                                                      'Brig Gen' THEN 0 WHEN 'Col' THEN 1 WHEN 'Lt Col' THEN 2 \
                       WHEN 'Maj' THEN 3 WHEN 'Capt' THEN 4 WHEN 'Lt' THEN 5 WHEN '2Lt' THEN 6 WHEN 'AWO' THEN 7 WHEN 'MWO' THEN 8 \
                       WHEN 'WO' THEN 9 WHEN 'SSgt' THEN 10 WHEN 'Sgt' THEN 11 WHEN 'Cpl' THEN 12 WHEN 'LCpl' THEN 13 \
                       WHEN 'Pte'  THEN 14 END, armynumber ASC", (search_string,))
                search_result = gender_search_result.fetchall()
                search_next_btn = tkinter.Button(per_info_frame, text='Next', bg="#93c47d", command=lambda: next_per(
                    'search_gender_btn'))
                search_next_btn.grid(row=6, column=3, sticky='e')
                search_prev_btn = tkinter.Button(per_info_frame, text='Back ', bg="#93c47d", command=lambda: prev_per(
                    'search_gender_btn'))
                search_prev_btn.grid(row=6, column=3, sticky='w')

        if search_result:
            clear_all_fields()
            clear_button.config(state='disabled')
            register_personnel_btn.config(state='disabled')
            go_forward_btn.config(state='disabled')
            go_back_btn.config(state='disabled')
            Serial_Number.set(search_result[i][0])
            Army_Number.set(search_result[i][1])
            Rank.set(search_result[i][2])
            First_Name.set(search_result[i][3])
            Last_Name.set(search_result[i][4])
            Gender_radio.set(search_result[i][5])
            Unit.set(search_result[i][6])
            Corps.set(search_result[i][7])
            Appointment.set(search_result[i][8])
            Deployment.set(search_result[i][9])
            State.set(search_result[i][10])
            Trade.set(search_result[i][11])
            address_entry.insert(INSERT, search_result[i][12])
            Marital_Status_radio.set(search_result[i][13])
            blob_image = search_result[i][14]
            blob = io.BytesIO(search_result[i][14])
            image = Image.open(blob)
            resized_photo = image.resize((300, 310))
            per_image = ImageTk.PhotoImage(resized_photo)
            img_label.config(image=per_image)
            img_label = per_image
        else:
            tkinter.messagebox.showinfo("Not Found", "Not found in DB")
        connection.close()
        return


# function to scroll to next record on returned search query
def next_per(button):
    global i
    if i == len(search_result) - 1:
        messagebox.showinfo("End", "Final Search Personnel")
    else:
        i += 1
        search_master(button)
    return


# function to scroll to previous record on returned search query
def prev_per(button):
    global i
    if i == 0:
        messagebox.showinfo("Start", "First Personnel")
    else:
        i -= 1
        search_master(button)
    return


# Frame to collect all required information
per_info_frame = tkinter.LabelFrame(frame, text="Personnel Information")
per_info_frame.grid(row=0, column=0)


# Labels and Entries
army_no_label = tkinter.Label(per_info_frame, text="Army Number")
army_no_label.grid(row=0, column=0)
Army_Number = StringVar()
army_no_entry = tkinter.Entry(per_info_frame, textvariable=Army_Number)
army_no_entry.grid(row=1, column=0)
rank_label = tkinter.Label(per_info_frame, text="Rank")
rank_label.grid(row=0, column=1)
Rank = StringVar()
rank_entry = ttk.Combobox(per_info_frame, values=["Brig Gen", "Col", "Lt Col", "Maj", "Capt", "Lt", "2Lt", "AWO", "MWO",
                                                  "WO", "SSgt", "Sgt", "Cpl", "LCpl", "Pte"], state="readonly", textvariable=Rank)
rank_entry.grid(row=1, column=1)
first_name_label = tkinter.Label(per_info_frame, text="First Name")
first_name_label.grid(row=0, column=2)
First_Name = StringVar()
first_name_entry = tkinter.Entry(per_info_frame, textvariable=First_Name)
first_name_entry.grid(row=1, column=2)
last_name_label = tkinter.Label(per_info_frame, text="Last Name")
last_name_label.grid(row=0, column=3)
Last_Name = StringVar()
last_name_entry = tkinter.Entry(per_info_frame, textvariable=Last_Name)
last_name_entry.grid(row=1, column=3)
unit_label = tkinter.Label(per_info_frame, text="Unit")
unit_label.grid(row=0, column=4)
Unit = StringVar()
unit_entry = tkinter.Entry(per_info_frame, textvariable=Unit)
unit_entry.grid(row=1, column=4)
gender_label = tkinter.Label(per_info_frame, text="Gender")
gender_label.grid(row=0, column=5)
Gender_radio = StringVar()
R1 = Radiobutton(per_info_frame, text="Male", variable=Gender_radio, value="Male")
R1.grid(row=1, column=5, sticky='w')
R2 = Radiobutton(per_info_frame, text="Female", variable=Gender_radio, value="Female")
R2.grid(row=2, column=5, sticky='w')
status_label = tkinter.Label(per_info_frame, text="Marital Status")
status_label.grid(row=0, column=6)
Marital_Status_radio = tkinter.StringVar()
R3 = Radiobutton(per_info_frame, text="Single", variable=Marital_Status_radio, value="Single")
R3.grid(row=1, column=6, sticky='w')
R4 = Radiobutton(per_info_frame, text="Married", variable=Marital_Status_radio, value="Married")
R4.grid(row=2, column=6, sticky='w')
R5 = Radiobutton(per_info_frame, text="Divorced", variable=Marital_Status_radio, value="Divorced")
R5.grid(row=3, column=6, sticky='w')
corps_label = tkinter.Label(per_info_frame, text="Corps")
corps_label.grid(row=2, column=0)
Corps = StringVar()
corps_entry = tkinter.Entry(per_info_frame, textvariable=Corps)
corps_entry.grid(row=3, column=0)
appointment_label = tkinter.Label(per_info_frame, text="Appointment")
appointment_label.grid(row=2, column=1)
Appointment = StringVar()
appointment_entry = tkinter.Entry(per_info_frame, textvariable=Appointment)
appointment_entry.grid(row=3, column=1)
deploy_label = tkinter.Label(per_info_frame, text="Deployment")
deploy_label.grid(row=2, column=2)
Deployment = StringVar()
deploy_entry = tkinter.Entry(per_info_frame, textvariable=Deployment)
deploy_entry.grid(row=3, column=2)
state_label = tkinter.Label(per_info_frame, text="State")
state_label.grid(row=2, column=3)
State = StringVar()
state_entry = ttk.Combobox(per_info_frame, values=["Abia", "Adamawa", "Akwa Ibom", "Anambra", "Bauchi", "Bayelsa",
                                                   "Benue", "Borno", "Cross River", "Delta", "Ebonyi", "Edo", "Ekiti",
                                                   "Enugu", "FCT - Abuja", "Gombe", "Imo", "Jigawa", "Kaduna", "Kano",
                                                   "Katsina", "Kebbi", "Kogi", "Kwara", "Lagos", "Nasarawa", "Niger",
                                                   "Ogun", "Ondo", "Osun", "Oyo", "Plateau", "Rivers", "Sokoto",
                                                   "Taraba", "Yobe", "Zamfara"], state="readonly", textvariable=State)
state_entry.grid(row=3, column=3)
trade_label = tkinter.Label(per_info_frame, text="Trade Class")
trade_label.grid(row=2, column=4)
Trade = StringVar()
trade_entry = tkinter.Entry(per_info_frame, textvariable=Trade)
trade_entry.grid(row=3, column=4)
address_label = tkinter.Label(per_info_frame, text="Accommodation Address")
address_label.grid(row=4, column=0, columnspan=2)
address_entry = tkinter.Text(per_info_frame, width=10, height=10, padx=3, pady=2, wrap=WORD)
address_entry.grid(row=5, column=0, columnspan=2, sticky="new")
per_img_frame = Frame(per_info_frame, bg="black", width=300, height=310, borderwidth=3, relief=GROOVE)
per_img_frame.grid(row=5, column=4, columnspan=2)
img = Image.open("fotos/passport.png")
resized_image = img.resize((300, 310))
photo = ImageTk.PhotoImage(resized_image)
img_label = Label(per_img_frame, image=photo)
img_label.place(x=0, y=0)
serial_label = tkinter.Label(per_info_frame, text="Serial No.", font="13")
serial_label.grid(row=4, column=6)
Serial_Number = IntVar()
serial_number_entry = tkinter.Entry(per_info_frame, textvariable=Serial_Number, font="13", state="readonly", width=5, bd=5,
                                    justify=RIGHT)
serial_number_entry.grid(row=5, column=6, sticky='n')


# Major Buttons
upload_image_btn = tkinter.Button(per_info_frame, text="Upload\nImage", command=upload_img, bg='lightblue')
upload_image_btn.grid(row=5, column=3, sticky='e')
clear_image_btn = tkinter.Button(per_info_frame, text="Clear\nImage", command=clear_image, bg='lightpink')
clear_image_btn.grid(row=5, column=3, sticky='se')
clear_button = tkinter.Button(per_info_frame, text="CLEAR ALL FIELDS", command=clear_all_fields, bg='#eb8080')
clear_button.grid(row=6, column=0, sticky='news', padx=20, pady=2)
register_personnel_btn = tkinter.Button(per_info_frame, text="REGISTER PERSONNEL", command=reg_per, bg='lightgreen')
register_personnel_btn.grid(row=7, column=0, sticky='news', padx=20, pady=2)
update_button = tkinter.Button(per_info_frame, text="UPDATE INFORMATION", bg='#a980eb', command=update_record)
update_button.grid(row=8, column=0, sticky='news', padx=20, pady=2)
return_button = tkinter.Button(per_info_frame, text="Go Back", command=go_back, bg="#ce7e00")
return_button.grid(row=9, column=0, sticky='news', padx=20, pady=2)
delete_personnel_btn = tkinter.Button(per_info_frame, text="REMOVE PERSONNEL", bg="#cc0000", command=delete_record)
delete_personnel_btn.grid(row=6, column=1, sticky='news')

# Peruse database buttons
restart_btn = tkinter.Button(per_info_frame, text="RETURN TO INPUT PERSONNEL", bg="#45818e", command=restart, font="Arial, 7")
restart_btn.grid(row=8, column=1)
peruse_db_btn = tkinter.Button(per_info_frame, text="PERUSE DATABASE", bg="#45818e", command=lambda: [peruse_database(), reset_i()])
peruse_db_btn.grid(row=9, column=1)
icon1 = PhotoImage(file="icons/arrow-left-bold-circle-outline.png")
icon2 = PhotoImage(file="icons/arrow-right-bold-circle-outline.png")
go_back_btn = tkinter.Button(per_info_frame, image=icon1, command=previous_personnel)
go_back_btn.grid(row=9, column=2, sticky='w')
go_forward_btn = tkinter.Button(per_info_frame, image=icon2, command=next_personnel)
go_forward_btn.grid(row=9, column=2, sticky='e')

# Search Widgets
find_label = tkinter.Label(per_info_frame, text='FIND:', font='bold 13')
find_label.grid(row=7, column=3, sticky='w')
search_string_entry = tkinter.Entry(per_info_frame, font='13')
search_string_entry.grid(row=8, column=3, sticky='w')
search_icon = PhotoImage(file="icons/search.png")
name_search_btn = tkinter.Button(per_info_frame, text="Search NAME", image=search_icon, bg="#a3c2c2", compound=RIGHT,
                                 command=search_database_name)
name_search_btn.grid(row=9, column=3, sticky='news')
army_number_search_btn = tkinter.Button(per_info_frame, text="Search ARMY NUMBER", image=search_icon, bg="#a3c2c2", compound=RIGHT,
                                        command=lambda: [reset_i(), search_master('search_army_no_btn')])
army_number_search_btn.grid(row=6, column=4, sticky='news')
rank_search_btn = tkinter.Button(per_info_frame, text="Search RANK", image=search_icon, bg="#a3c2c2", compound=RIGHT,
                                 command=lambda: [reset_i(), search_master('search_rank_btn')])
rank_search_btn.grid(row=7, column=4, sticky='news')
deployment_search_btn = tkinter.Button(per_info_frame, text="Search DEPLOYMENT", image=search_icon, bg="#a3c2c2", compound=RIGHT,
                                       command=lambda: [reset_i(), search_master('search_deployment_btn')])
deployment_search_btn.grid(row=8, column=4, sticky='news')
state_search_btn = tkinter.Button(per_info_frame, text="Search STATE", image=search_icon, bg="#a3c2c2", compound=RIGHT,
                                  command=lambda: [reset_i(), search_master('search_state_btn')])
state_search_btn.grid(row=9, column=4, sticky='news')
trade_search_btn = tkinter.Button(per_info_frame, text="Search TRADE", image=search_icon, bg="#a3c2c2", compound=RIGHT,
                                  command=lambda: [reset_i(), search_master('search_trade_btn')])
trade_search_btn.grid(row=6, column=5, columnspan=2, sticky='news')
corps_search_btn = tkinter.Button(per_info_frame, text="Search CORPS", image=search_icon, bg="#a3c2c2", compound=RIGHT,
                                  command=lambda: [reset_i(), search_master('search_corps_btn')])
corps_search_btn.grid(row=7, column=5, columnspan=2, sticky='news')
appointment_search_btn = tkinter.Button(per_info_frame, text="Search APPOINTMENT", image=search_icon, bg="#a3c2c2", compound=RIGHT,
                                        command=lambda: [reset_i(), search_master('search_appointment_btn')])
appointment_search_btn.grid(row=8, column=5, columnspan=2, sticky='news')
gender_search_btn = tkinter.Button(per_info_frame, text="Search GENDER", image=search_icon, bg="#a3c2c2", compound=RIGHT,
                                   command=lambda: [reset_i(), search_master('search_gender_btn')])
gender_search_btn.grid(row=9, column=5, columnspan=2, sticky='news')

for widget in per_info_frame.winfo_children():
    widget.grid_configure(padx=10, pady=2)

window.mainloop()
